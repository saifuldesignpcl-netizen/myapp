import flet as ft
from datetime import date
import openpyxl
import os
import re
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors as pdf_colors

MASTER_DATA_FILENAME = "cable_master.xlsx"
COLOR_OPTIONS = ["Red", "Black", "Yellow", "Blue", "Green", "Gray", "White"]
UNIT_OPTIONS = ["Coil", "Meter"]
COIL_MULTIPLIER = 100

DEFAULT_TERMS = [
    ["Delivery",
     "Delivery will be made to your project site at our cost within 7 working days from the receiving of your work order."],
    ["Payment", "100% advanced with work order by DD / P.O / Cheque / Cash / EFT in favor of the company."],
    ["Offer Validity", "15 (Fifteen) working days from the date of issue of this offer."],
    ["Quality Inspection", "Pre-shipment inspection is the final inspection for quality and any technical argument."],
    ["Tolerance", "(+/-) 2% should be allowed."],
    ["Mode of Modification",
     "Purchase order once received will be treated as firm and final. No deviation accepted after PO."],
]


def clean_str(val):
    if val is None:
        return ""
    # Non-breaking space (\xa0) এবং অতিরিক্ত স্পেস রিমুভ করার জন্য
    txt = str(val).replace('\xa0', ' ')
    return re.sub(r'\s+', ' ', txt).strip()


def number_to_words(num: float) -> str:
    num = int(round(num))
    if num == 0:
        return "Zero Taka Only"

    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight",
            "Nine", "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen",
            "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy",
            "Eighty", "Ninety"]

    def two_digit(n):
        if n < 20:
            return ones[n]
        return tens[n // 10] + (" " + ones[n % 10] if n % 10 else "")

    def three_digit(n):
        if n < 100:
            return two_digit(n)
        rest = n % 100
        return ones[n // 100] + " Hundred" + (" " + two_digit(rest) if rest else "")

    crore, num = divmod(num, 10_000_000)
    lakh, num = divmod(num, 100_000)
    thousand, rest = divmod(num, 1000)

    parts = []
    if crore:
        parts.append(three_digit(crore) + " Crore")
    if lakh:
        parts.append(three_digit(lakh) + " Lakh")
    if thousand:
        parts.append(three_digit(thousand) + " Thousand")
    if rest:
        parts.append(three_digit(rest))

    return " ".join(parts) + " Taka Only"


class MasterData:
    COLUMN_ALIASES = {
        "type": ["type", "product type", "cable type", "spec", "product specification"],
        "size": ["size", "cable size", "sq mm", "sqmm"],
        "price": ["price", "net price", "unit price", "rate", "net price in bdt"],
    }

    def __init__(self):
        self.rows = []

    def load(self, path):
        if not os.path.exists(path):
            return
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        header_row = [clean_str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_index = {}
        for field, aliases in self.COLUMN_ALIASES.items():
            for i, h in enumerate(header_row):
                if h.lower() in aliases:
                    col_index[field] = i
                    break

        if "type" not in col_index or "size" not in col_index:
            return

        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            def get(field, default=""):
                idx = col_index.get(field)
                if idx is None or idx >= len(r) or r[idx] is None:
                    return default
                return r[idx]

            t = clean_str(get("type"))
            s = clean_str(get("size"))
            if not t and not s:
                continue
            try:
                p = float(get("price", 0))
            except (TypeError, ValueError):
                p = 0.0

            rows.append({"type": t, "size": s, "price": p})

        self.rows = rows

    def unique(self, field):
        return sorted(list({row[field] for row in self.rows if row[field]}))

    def sizes_for_type(self, ctype):
        if not ctype:
            return []
        target = clean_str(ctype).lower()
        sizes = []
        for row in self.rows:
            if clean_str(row["type"]).lower() == target and row["size"]:
                if row["size"] not in sizes:
                    sizes.append(row["size"])
        return sorted(sizes)

    def find_price(self, ctype, csize):
        if not ctype or not csize:
            return None
        target_type = clean_str(ctype).lower()
        target_size = clean_str(csize).lower()
        for row in self.rows:
            if clean_str(row["type"]).lower() == target_type and clean_str(row["size"]).lower() == target_size:
                return row["price"]
        return None


def generate_pdf(filename, info, items, grand_total, in_words, terms):
    doc = SimpleDocTemplate(filename, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    styles = getSampleStyleSheet()

    # Title & Header
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, leading=20,
                                 textColor=pdf_colors.HexColor('#B71C1C'))
    story.append(Paragraph(info['company_name'], title_style))
    story.append(Paragraph(info['company_addr'], styles['Normal']))
    story.append(Spacer(1, 10))

    meta_text = f"<b>Ref:</b> {info['ref_no']} | <b>Date:</b> {info['doc_date']}<br/><b>Client:</b> {info['client_name']}<br/><b>Subject:</b> {info['subject']}"
    story.append(Paragraph(meta_text, styles['Normal']))
    story.append(Spacer(1, 15))

    # Items Table
    table_data = [["SL", "Type", "Size", "Color", "Unit", "Qty", "Price", "Total"]]
    for idx, item in enumerate(items, 1):
        tot = item['qty'] * item['price']
        table_data.append([
            str(idx), item['type'], item['size'], item['color'], item['unit'],
            str(item['qty']), f"{item['price']:.2f}", f"{tot:.2f}"
        ])

    item_table = Table(table_data, colWidths=[25, 100, 70, 50, 45, 35, 65, 75])
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), pdf_colors.HexColor('#B71C1C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), pdf_colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, pdf_colors.grey),
    ]))
    story.append(item_table)
    story.append(Spacer(1, 15))

    # Total & Words
    story.append(Paragraph(f"<b>Grand Total: {grand_total:,.2f} BDT</b>", styles['Heading2']))
    story.append(Paragraph(f"<b>In Words:</b> {in_words}", styles['Italic']))
    story.append(Spacer(1, 15))

    # Terms
    story.append(Paragraph("<b>Terms & Conditions:</b>", styles['Heading3']))
    terms_data = [[t[0], t[1]] for t in terms]
    terms_table = Table(terms_data, colWidths=[120, 350])
    terms_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, pdf_colors.lightgrey),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    story.append(terms_table)

    doc.build(story)


def main(page: ft.Page):
    page.title = "Cable Quotation Generator"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.scroll = ft.ScrollMode.AUTO
    page.padding = 16

    master = MasterData()
    master.load(MASTER_DATA_FILENAME)

    items = []
    terms = [list(t) for t in DEFAULT_TERMS]
    base_price = [None]

    # --- Inputs ---
    company_name = ft.TextField(label="Company Name", value="ELCO WIRES AND CABLES LIMITED", dense=True)
    company_addr = ft.TextField(label="Company Address", value="102, Shukrabad, Mirpur Road, Dhaka", dense=True)
    ref_no = ft.TextField(label="Reference No", value=f"QT/{date.today().strftime('%d%m%Y')}", dense=True)
    doc_date = ft.TextField(label="Date", value=date.today().strftime('%d/%m/%Y'), dense=True)
    client_name = ft.TextField(label="Client Name", value="Client Company Limited", dense=True)
    client_addr = ft.TextField(label="Client Address", value="Client Address", dense=True)
    subject_line = ft.TextField(label="Subject", value="Price Offer for Electrical Cables.", dense=True)

    # Dropdowns Setup
    types_list = master.unique("type")

    type_dropdown = ft.Dropdown(
        label="Type",
        options=[ft.dropdown.Option(key=t, text=t) for t in types_list],
        dense=True,
        expand=True
    )
    size_dropdown = ft.Dropdown(
        label="Size",
        options=[],
        dense=True,
        expand=True
    )
    color_dropdown = ft.Dropdown(
        label="Color",
        options=[ft.dropdown.Option(key=c, text=c) for c in [""] + COLOR_OPTIONS],
        value="",
        dense=True,
        expand=True
    )
    unit_dropdown = ft.Dropdown(
        label="Unit",
        options=[ft.dropdown.Option(key=u, text=u) for u in UNIT_OPTIONS],
        value="Meter",
        dense=True,
        expand=True
    )
    qty_field = ft.TextField(label="Quantity", value="1", keyboard_type=ft.KeyboardType.NUMBER, dense=True, expand=True)
    price_field = ft.TextField(label="Net Price (BDT)", value="0.00", keyboard_type=ft.KeyboardType.NUMBER, dense=True,
                               expand=True)

    net_total_text = ft.Text("0.00", weight=ft.FontWeight.BOLD)
    discount_field = ft.TextField(label="Special Discount (%)", value="0.0", keyboard_type=ft.KeyboardType.NUMBER,
                                  dense=True, width=150)
    grand_total_text = ft.Text("0.00", size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_700)
    words_text = ft.Text("Zero Taka Only", italic=True)

    items_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("SL")),
            ft.DataColumn(ft.Text("Type")),
            ft.DataColumn(ft.Text("Size")),
            ft.DataColumn(ft.Text("Color")),
            ft.DataColumn(ft.Text("Unit")),
            ft.DataColumn(ft.Text("Qty")),
            ft.DataColumn(ft.Text("Price")),
            ft.DataColumn(ft.Text("Total")),
            ft.DataColumn(ft.Text("Action")),
        ],
        rows=[]
    )

    terms_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("Label")),
            ft.DataColumn(ft.Text("Detail")),
        ],
        rows=[]
    )

    def recalc():
        net = sum(item["qty"] * item["price"] for item in items)
        try:
            disc_pct = float(discount_field.value or 0)
        except ValueError:
            disc_pct = 0.0

        discount_val = net * (disc_pct / 100.0)
        grand = max(net - discount_val, 0.0)

        net_total_text.value = f"{net:,.2f} BDT"
        grand_total_text.value = f"{grand:,.2f} BDT"
        words_text.value = number_to_words(grand)
        page.update()

    def update_price():
        t = type_dropdown.value
        s = size_dropdown.value
        if t and s:
            p = master.find_price(t, s)
            base_price[0] = p
            if p is not None:
                if unit_dropdown.value == "Coil":
                    price_field.value = f"{p * COIL_MULTIPLIER:.2f}"
                else:
                    price_field.value = f"{p:.2f}"
            else:
                price_field.value = "0.00"
        else:
            price_field.value = "0.00"

    def update_sizes_for_selected_type(selected_type):
        sizes = master.sizes_for_type(selected_type)

        # অপশন ক্লিয়ার করে নতুন অপশন সেট করা
        size_dropdown.options.clear()
        for s in sizes:
            size_dropdown.options.append(ft.dropdown.Option(key=s, text=s))

        if sizes:
            size_dropdown.value = sizes[0]
        else:
            size_dropdown.value = None

        update_price()
        page.update()  # পুরো পেজ আপডেট দিয়ে UI সিঙ্ক নিশ্চিত করা হলো

    def on_type_change(e):
        selected_type = e.control.value
        unit_dropdown.value = "Meter"
        update_sizes_for_selected_type(selected_type)

    def on_size_change(e):
        unit_dropdown.value = "Meter"
        update_price()
        page.update()

    def on_unit_change(e):
        if base_price[0] is not None:
            if unit_dropdown.value == "Coil":
                price_field.value = f"{base_price[0] * COIL_MULTIPLIER:.2f}"
            else:
                price_field.value = f"{base_price[0]:.2f}"
            page.update()

    # Flet's Dropdown widget renamed its "value picked" event from
    # on_change to on_select in recent versions (TextField still uses
    # on_change, unaffected) - using on_change here silently did nothing,
    # which is why choosing a Type/Size/Unit never updated anything.
    type_dropdown.on_select = on_type_change
    size_dropdown.on_select = on_size_change
    unit_dropdown.on_select = on_unit_change
    discount_field.on_change = lambda e: recalc()

    def refresh_items_table():
        items_table.rows.clear()
        for idx, item in enumerate(items):
            total = item["qty"] * item["price"]

            def delete_row(e, index=idx):
                items.pop(index)
                refresh_items_table()
                recalc()

            items_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(idx + 1))),
                        ft.DataCell(ft.Text(item["type"])),
                        ft.DataCell(ft.Text(item["size"])),
                        ft.DataCell(ft.Text(item["color"])),
                        ft.DataCell(ft.Text(item["unit"])),
                        ft.DataCell(ft.Text(str(item["qty"]))),
                        ft.DataCell(ft.Text(f"{item['price']:.2f}")),
                        ft.DataCell(ft.Text(f"{total:.2f}")),
                        ft.DataCell(
                            ft.IconButton(
                                icon=ft.Icons.DELETE,
                                icon_color=ft.Colors.RED_400,
                                on_click=delete_row
                            )
                        ),
                    ]
                )
            )
        page.update()

    def refresh_terms_table():
        terms_table.rows.clear()
        for term in terms:
            terms_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(term[0])),
                        ft.DataCell(ft.Text(term[1])),
                    ]
                )
            )
        page.update()

    def show_snack(text):
        # page.open(snack) doesn't exist in this Flet version - the
        # correct pattern is: add the SnackBar to page.overlay, set its
        # own .open flag, then page.update().
        snack = ft.SnackBar(ft.Text(text))
        page.overlay.append(snack)
        snack.open = True
        page.update()

    def add_item_click(e):
        if not type_dropdown.value or not size_dropdown.value:
            show_snack("Please select Type and Size!")
            return

        try:
            qty = int(qty_field.value)
            price = float(price_field.value)
        except ValueError:
            return

        items.append({
            "type": type_dropdown.value,
            "size": size_dropdown.value,
            "color": color_dropdown.value or "",
            "unit": unit_dropdown.value,
            "qty": qty,
            "price": price,
        })
        refresh_items_table()
        recalc()

    def export_pdf_click(e):
        if not items:
            show_snack("No items to export! Please add items first.")
            return

        try:
            info = {
                "company_name": company_name.value,
                "company_addr": company_addr.value,
                "ref_no": ref_no.value,
                "doc_date": doc_date.value,
                "client_name": client_name.value,
                "subject": subject_line.value
            }
            net = sum(item["qty"] * item["price"] for item in items)
            disc_pct = float(discount_field.value or 0)
            grand = max(net - (net * disc_pct / 100.0), 0.0)

            output_pdf = "Cable_Quotation.pdf"
            generate_pdf(output_pdf, info, items, grand, words_text.value, terms)

            show_snack(f"PDF Exported Successfully as '{output_pdf}'!")
        except Exception as ex:
            show_snack(f"Error generating PDF: {str(ex)}")

    # --- UI Layout ---
    page.add(
        ft.AppBar(
            title=ft.Text("Cable Quotation App"),
            bgcolor=ft.Colors.RED_800,
            color=ft.Colors.WHITE
        ),
        ft.Card(
            content=ft.Container(
                content=ft.Column([
                    ft.Text("Quotation Info", size=16, weight=ft.FontWeight.BOLD),
                    company_name,
                    company_addr,
                    ft.Row([ref_no, doc_date]),
                    client_name,
                    client_addr,
                    subject_line
                ]),
                padding=12
            )
        ),
        ft.Card(
            content=ft.Container(
                content=ft.Column([
                    ft.Text("Add Cable Item", size=16, weight=ft.FontWeight.BOLD),
                    ft.Row([type_dropdown, size_dropdown]),
                    ft.Row([color_dropdown, unit_dropdown]),
                    ft.Row([qty_field, price_field]),
                    ft.Button(
                        "Add to Quotation",
                        icon=ft.Icons.ADD,
                        style=ft.ButtonStyle(bgcolor=ft.Colors.RED_700, color=ft.Colors.WHITE),
                        on_click=add_item_click
                    )
                ]),
                padding=12
            )
        ),
        ft.Card(
            content=ft.Container(
                content=ft.Column([
                    ft.Text("Quotation Items", size=16, weight=ft.FontWeight.BOLD),
                    ft.Row([items_table], scroll=ft.ScrollMode.ALWAYS)
                ]),
                padding=12
            )
        ),
        ft.Card(
            content=ft.Container(
                content=ft.Column([
                    ft.Text("Totals", size=16, weight=ft.FontWeight.BOLD),
                    ft.Row([ft.Text("Net Total: "), net_total_text]),
                    discount_field,
                    ft.Row([ft.Text("Grand Total: "), grand_total_text]),
                    ft.Row([ft.Text("In Words: "), words_text])
                ]),
                padding=12
            )
        ),
        ft.Card(
            content=ft.Container(
                content=ft.Column([
                    ft.Text("Terms & Conditions", size=16, weight=ft.FontWeight.BOLD),
                    ft.Row([terms_table], scroll=ft.ScrollMode.ALWAYS)
                ]),
                padding=12
            )
        ),
        ft.Container(
            content=ft.Button(
                "Export PDF",
                icon=ft.Icons.PICTURE_AS_PDF,
                style=ft.ButtonStyle(bgcolor=ft.Colors.RED_800, color=ft.Colors.WHITE),
                height=50,
                on_click=export_pdf_click
            ),
            padding=ft.Padding.only(top=10, bottom=20),
            alignment=ft.Alignment(0, 0)
        )
    )

    # --- Initial Selection ---
    if types_list:
        type_dropdown.value = types_list[0]
        update_sizes_for_selected_type(types_list[0])

    refresh_terms_table()


if __name__ == "__main__":
    ft.run(main)